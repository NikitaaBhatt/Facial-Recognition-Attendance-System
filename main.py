import os
import pickle
import numpy as np
import cv2
import face_recognition
from datetime import datetime
import openpyxl

# Load the encoding file
print("Loading Encode File ...")
file = open('EncodeFile.p', 'rb')
encodeListKnownWithIds = pickle.load(file)
file.close()
encodeListKnown, employeeIds = encodeListKnownWithIds
print("Encode File Loaded")

# Initialize the webcam
cap = cv2.VideoCapture(0)
cap.set(3, 640)  # Width
cap.set(4, 480)  # Height

# Load the background and mode images
imgBackground = cv2.imread('Resources/background.png')
folderModePath = 'Resources/Modes'
modePathList = os.listdir(folderModePath)
imgModeList = [cv2.imread(os.path.join(folderModePath, path)) for path in modePathList]

modeType = 0
counter = 0
employee_id = -1
imgEmployee = []

while True:
    success, img = cap.read()

    # Resize and convert the image for face recognition
    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

    # Update the background image with the live webcam feed
    # Adjust the webcam placement to match the face in the background image
    imgBackground[160:160 + 480, 55:55 + 640] = img

    # Resize and place the current mode image
    mode_image_resized = cv2.resize(imgModeList[modeType], (414, 633))
    imgBackground[44:44 + 633, 808:808 + 414] = mode_image_resized

    faceCurFrame = face_recognition.face_locations(imgS)
    encodeCurFrame = face_recognition.face_encodings(imgS, faceCurFrame)

    if faceCurFrame:
        for encodeFace, faceLoc in zip(encodeCurFrame, faceCurFrame):
            matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
            faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
            matchIndex = np.argmin(faceDis)

            if matches[matchIndex]:
                employee_id = employeeIds[matchIndex]

                if counter == 0:
                    modeType = 1  # Switch to the second mode (showing student details)
                    counter = 1

                    # Load the attendance sheet
                    workbook = openpyxl.load_workbook("Attendance.xlsx")
                    sheet = workbook.active
                    found = False

                    # Check and update attendance
                    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                        if row[0].value == employee_id:
                            found = True
                            last_attendance_time = row[-1].value
                            current_time = datetime.now()

                            # Check if attendance can be marked
                            if last_attendance_time is None or (current_time - last_attendance_time).days > 0:
                                row[2].value += 1  # Increment attendance
                                row[-1].value = current_time  # Update last attendance time
                                modeType = 2  # Marked attendance mode
                            else:
                                modeType = 3  # Already marked mode
                            break

                    if not found:
                        modeType = 3  # Not found in records

                    workbook.save("Attendance.xlsx")
                    counter = 1

        if counter > 0:
            counter += 1

            if 10 < counter < 20:
                modeType = 2  # Display student details
                cv2.putText(imgBackground, f"ID: {employee_id}", (950, 400), cv2.FONT_HERSHEY_SIMPLEX, 0.7,
                            (255, 255, 255), 2)
                cv2.putText(imgBackground, "Attendance Marked!", (950, 450), cv2.FONT_HERSHEY_SIMPLEX, 0.7,
                            (255, 255, 255), 2)

            if counter >= 20:
                counter = 0
                modeType = 0  # Reset to the default mode

    else:
        modeType = 0
        counter = 0

    # Display the GUI window
    cv2.imshow("Face Attendance", imgBackground)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()